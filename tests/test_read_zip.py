from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook

from tests.paths import ZIP_FILE


def test_count_files():
    with ZipFile(ZIP_FILE) as zf:
        zip_contents = zf.namelist()
        num_files = len(zip_contents)
    assert num_files == 3


def test_read_pdf_file():
    with ZipFile(ZIP_FILE) as zf:
        with zf.open('file_1.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            text = reader.pages[0].extract_text()
    assert 'Ответ на запрос АФМ' in text


def test_read_xlsx():
    with ZipFile(ZIP_FILE) as zf:
        with zf.open('file_2.xlsx') as xlsx_file:
            table = load_workbook(xlsx_file)
            sheet = table.active

    assert sheet.cell(row=1, column=1).value == 'Журнал запросов/ответов из АФМ'


def test_csv_read():
    with ZipFile(ZIP_FILE) as zf:
        with zf.open('file_3.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')

    assert 'test' in content
